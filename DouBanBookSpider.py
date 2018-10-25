#!/usr/bin
#-*- coding: UTF-8 -*-

import requests
import re
import time
import numpy as np
from bs4 import BeautifulSoup
from openpyxl import Workbook
from pymongo import MongoClient

#备选Agents
hds=[{'User-Agent':'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-US; rv:1.9.1.6) Gecko/20091201 Firefox/3.5.6'},\
{'User-Agent':'Mozilla/5.0 (Windows NT 6.2) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.12 Safari/535.11'},\
{'User-Agent': 'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.2; Trident/6.0)'}]

###定义MongoDB数据库IP、端口和数据库名
client = MongoClient('127.0.0.1', 27017)
db = client.douban_book

###指定集合名称
coll = db['book_by_tag']

def book_spider(book_tag):
    page_count=0;
    book_list=[]
    try_times=0
    
    while(1):
        url='http://www.douban.com/tag/'+book_tag+'/book?start='+str(page_count*15)
        # time.sleep(np.random.rand()*5)
        
        #Last Version
        try:
            req = requests.get(url, headers=hds[page_count%len(hds)])
            plain_text = req.text
        except:
            continue
        
        soup = BeautifulSoup(plain_text,"lxml")
        list_soup = soup.find('div', {'class': 'mod book-list'})
        
        #控制采集深度
        try_times+=1;
        if list_soup==None and try_times<10:
            continue
        elif list_soup==None or len(list_soup)<=1:
            break # Break when no informatoin got after 200 times requesting
        
        # for book_info in list_soup.findAll('dd'):
        #     title = book_info.find('a', {'class':'title'}).string.strip()
        #     desc = book_info.find('div', {'class':'desc'}).string.strip()
        #     desc_list = desc.split('/')
        #     book_url = book_info.find('a', {'class':'title'}).get('href')

        for book_info_dl in list_soup.findAll('dl'):
        	#图片信息
        	img_info = book_info_dl.find('dt').find('a')
        	#图片链接
        	img_url = img_info.find('img').get('src')

        	#图书信息
        	book_info = book_info_dl.find('dd')
        	#书名
        	title = book_info.find('a', {'class':'title'}).string.strip()
        	#
        	desc = book_info.find('div', {'class':'desc'}).string.strip()
        	desc_list = desc.split('/')
        	#图书链接
        	book_url = book_info.find('a', {'class':'title'}).get('href')

        	#作者/译者
        	try:
        		author_info = '/'.join(desc_list[0:-3]).strip()
        	except:
        		author_info ='暂无'
        	
        	#出版信息
        	try:
        		pub_info = '/'.join(desc_list[-3:]).strip()
        	except:
        		pub_info = '暂无'
        	#评分
        	try:
        		rating = book_info.find('span', {'class':'rating_nums'}).string.strip()
        	except:
        		rating='0.0'
        	
        	#评价人数,页数,ISBN
        	pub_house,pub_year,price,people_num,page_num,isbn = get_detail_book_info(book_url)

        	#获取采集时间戳
        	data_time = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
        	#print(data_time)

        	book_list.append([title,rating,people_num,author_info,pub_house,pub_year,price,page_num,isbn,book_url,img_url,data_time])
        	book_info_to_mongodb(book_tag,title,rating,people_num,author_info,pub_house,pub_year,price,page_num,isbn,book_url,img_url,data_time)
        	try_times=0 #set 0 when got valid information
        page_count+=1
        print('Downloading Information From Page '+str(page_count))


def get_detail_book_info(url):
    #url='http://book.douban.com/subject/6082808/?from=tag_all' # For Test
    try:
        req = requests.get(url, headers=hds[np.random.randint(0,len(hds))])
        plain_text=req.text   
    except:
        print('someting goes wrong!')

    soup = BeautifulSoup(plain_text,"lxml")
    try:
    	people_num=soup.find('div',{'class':'rating_sum'}).findAll('span')[1].string.strip()
    except:
    	people_num = '0'

    detail_book_info = soup.find('div',{'id':'info'})
    # i = 1
    # for string in detail_book_info.stripped_strings:
    # 	print(str(i)+':   '+str(string.strip()))
    # 	i += 1
    
    #通过字符串搜索含有"ISBN"的tag,通过next_sibling查询兄弟节点
    try:
    	isbn = detail_book_info.find('span',string=re.compile('ISBN')).next_sibling.string.strip()
    except:
    	isbn = 'null'

    #通过字符串搜索含有"页数"的tag,通过next_sibling查询兄弟节点
    try:
    	page_num = detail_book_info.find('span',string=re.compile('页数')).next_sibling.string.strip()
    except:
    	page_num = 'null'

    #通过字符串搜索含有"出版社"的tag,通过next_sibling查询兄弟节点
    try:
    	pub_house = detail_book_info.find('span',string=re.compile('出版社')).next_sibling.string.strip()
    except:
    	pub_house = 'null'

	#通过字符串搜索含有"出版年"的tag,通过next_sibling查询兄弟节点
    try:
    	pub_year = detail_book_info.find('span',string=re.compile('出版年')).next_sibling.string.strip()
    except:
    	pub_year = 'null'

    #通过字符串搜索含有"定价"的tag,通过next_sibling查询兄弟节点
    try:
    	price = detail_book_info.find('span',string=re.compile('定价')).next_sibling.string.strip()
    except:
    	price = 'null'

    return pub_house,pub_year,price,people_num,page_num,isbn


def do_spider(book_tag_lists):
    book_lists=[]
    for book_tag in book_tag_lists:
    	print('Downloading Information From Tag '+book_tag)
    	book_spider(book_tag)
# def print_book_lists_excel(book_lists,book_tag_lists):
#     wb=Workbook(write_only=True)
#     ws=[]
#     for i in range(len(book_tag_lists)):
#         ws.append(wb.create_sheet(title=book_tag_lists[i])) #utf8->unicode
#     for i in range(len(book_tag_lists)): 
#         ws[i].append(['标签','序号','书名','评分','评价人数','作者/译者','出版社','出版年','定价','页数','ISBN','图书链接','图片链接','采集时间'])
#         count=1
#         for bl in book_lists[i]:
#             ws[i].append([book_tag_lists[i],count,bl[0],float(bl[1]),int(bl[2]),bl[3],bl[4],bl[5],bl[6],bl[7],bl[8],bl[9],bl[10],bl[11]])
#             count+=1
#     save_path='book_list'
#     for i in range(len(book_tag_lists)):
#         save_path+=('-'+book_tag_lists[i])
#     save_path+='.xlsx'
#     wb.save(save_path)

#将采集到的书籍信息逐本插入到mongodb
def book_info_to_mongodb(book_tag,title,rating,people_num,author_info,pub_house,pub_year,price,page_num,isbn,book_url,img_url,data_time):
	coll.insert_one({'book_tag':book_tag,'title':title,'rating':rating,'people_num':people_num,\
		'author_info':author_info,'pub_house':pub_house,'pub_year':pub_year,'price':price,\
		'page_num':page_num,'isbn':isbn,'book_url':book_url,'img_url':img_url,'data_time':data_time})


if __name__=='__main__':
    #book_tag_lists = ['心理','判断与决策','算法','数据结构','经济','历史']
    #book_tag_lists = ['传记','哲学','编程','创业','理财','社会学','佛教']
    #book_tag_lists = ['思想','科技','科学','web','股票','爱情','两性']
    #book_tag_lists = ['计算机','机器学习','linux','android','数据库','互联网']
    #book_tag_lists = ['数学']
    #book_tag_lists = ['摄影','设计','音乐','旅行','教育','成长','情感','育儿','健康','养生']
    #book_tag_lists = ['商业','理财','管理']  
    #book_tag_lists = ['名著']
    #book_tag_lists = ['科普','经典','生活','心灵','文学']
    #book_tag_lists = ['科幻','思维','金融']
    book_tag_lists = ['个人管理','时间管理','投资','文化','宗教']
    # print('step 001')
    book_lists=do_spider(book_tag_lists)
    # print('step 002')
    # print_book_lists_excel(book_lists,book_tag_lists)

# book_tag_lists = ['个人管理']
# print('step 001')
# book_lists=do_spider(book_tag_lists)
# print('step 002')
# print_book_lists_excel(book_lists,book_tag_lists)
